from collections import defaultdict
import os
import re
import openpyxl
import datetime
import tempfile
from django.db import connection
from asgiref.sync import async_to_sync, sync_to_async
import plotly.graph_objects as go
import pandas as pd


async def get_scores(excel):
    
    exam_rec = await sync_to_async(get_db_exam)()
    recs = []
    
    if exam_rec and 'ID' in exam_rec:
        exam_id = exam_rec['ID'] 
        recs = await sync_to_async(get_db_recs)(exam_id)
        print(f"Found exam ID: {exam_id}. Fetched {len(recs)} result records.")
    else:
        print("No eligible exam record found in HSP ExamListTable.")

    excel_out = await sync_to_async(get_excel)(exam_rec, recs)
    score_text = await sync_to_async(get_text)(recs)
    chart_out = await sync_to_async(get_chart)(exam_rec)
    excel_out.append(score_text)
    excel_out.append(chart_out)
    
    return excel_out

def get_db_exam():
    with connection.cursor() as cursor:
        cursor.execute("SELECT * FROM HSPExamListTable WHERE DATEADD(Week,1 , ExamDate) > GETDATE()")
        # Get column names immediately after execution
        columns = [column[0] for column in cursor.description]

        # Use fetchone() for the single expected record
        record = cursor.fetchone()

        if record:
            # Create a dictionary by zipping columns and the record values
            return dict(zip(columns, record))
        else:
            # Return None if no record is found
            return None
        
def get_db_recs(exam_id):
    with connection.cursor() as cursor:
        sql = "SELECT * FROM HSPExamNotPassedFunction(%s) Order By ODID, GID, Pos, ID"
        cursor.execute(sql, [exam_id])
        recs = [dict(zip([column[0] for column in cursor.description], record)) for record in cursor.fetchall()]
    return recs


def get_text(recs):
    if recs:
        grouped = defaultdict(list)

        # Group by Dept
        for rec in recs:
            dept = rec.get("Dept", "Unknown")
            grouped[dept].append(rec)
            
        output_lines = []
        for dept, items in grouped.items():
            num = len(items)
            output_lines.append(f"--- Department: {dept} ({num}) ---")
            for r in items:
                grp = r.get("Grp") or "Unknown"
                name = r.get("Name") or "Unknown"
                eid = r.get("ID") or ""
                score = r.get("Score") or "0"
                output_lines.append(f"{grp} {name} ({eid}) - {score}")
            output_lines.append("")  # Add blank line between departments
            

        formatted_text = "\n".join(output_lines)
        return formatted_text
    else:
        return "No records found."
    


def get_excel(exam_rec, recs):

    wb = openpyxl.load_workbook('apis/Telegram/Exam/assets/Exam_Template.xlsx')
    ws = wb.active # Or specify the sheet name
    
    exam_name = exam_rec.get('ExamName')
    
    exam_title = f"AUS church(Melbourne) HSP Exam ({exam_name}) Result"
    
    ws["A1"] = exam_title
    
    with connection.cursor() as cursor:
        sql = "SELECT * FROM HSPExamResultsTable WHERE ExamID = %s"
        cursor.execute(sql, [exam_rec.get('ID')])
        erecs = [dict(zip([column[0] for column in cursor.description], record)) for record in cursor.fetchall()]
    
    rec_map = {
        str(rec["EVID"]).strip(): rec
        for rec in erecs
        if rec.get("EVID") is not None
    }
    last_row = ws.max_row

    for row in range(6, last_row + 1):
        student_id = ws.cell(row=row, column=7).value  # Column F = 6th col
        if not student_id:
            continue

        student_id = str(student_id).strip()
        if student_id in rec_map:
            record = rec_map[student_id]
            score = record.get("Score")
            reason = record.get("ReasonNotDone")

            # Example: write score to column H (8th col) and reason to I (9th col)
            ws.cell(row=row, column=11).value = score
            ws.cell(row=row, column=12).value = reason
        else:
            ws.cell(row=row, column=11).value = "0"
            ws.cell(row=row, column=12).value = "No submission"
            
    
    pattern = r"[:.,]"
    replacement = "_"

    cleaned_text = re.sub(pattern, replacement, exam_name)
    
    # Get the current year
    current_year = datetime.datetime.now().year # Result: 2025
    start_year = 1984
    scj_yr = current_year - start_year + 1

    filename = f"HSP Exam {cleaned_text} Score_ReportEN.xlsx"
    
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)

    return [tmp.name, exam_title, filename]


def get_chart(exam_rec):
    exam_id = exam_rec['ID'] 
    exam_name = exam_rec['ExamName'] 
    with connection.cursor() as cursor:
        sql = "SELECT * FROM HSPExamPassRateFunction(%s) Order By ODID"
        cursor.execute(sql, [exam_id])
        recs = [dict(zip([column[0] for column in cursor.description], record)) for record in cursor.fetchall()]

    df = pd.DataFrame(recs)

    fig = go.Figure()

    # Stacked Bars
    fig.add_trace(go.Bar(
        x=df["Dept"],
        y=df["Passed"],
        name="Passed",
        marker=dict(
            color=df["PassRate"],          # numeric values
            colorscale=[          # ✅ custom gradient
                [0, "#A8E6CF"],   # light green
                [1, "#05668D"]    # dark teal
            ],       # built-in gradient
            showscale=False              # hide color bar if you don’t want it
        ),
        text=df["Passed"],
        textposition="inside",
        textfont=dict(color="white", size=12)
    ))

    # Failed bars with gradient based on value
    fig.add_trace(go.Bar(
        x=df["Dept"],
        y=df["Failed"],
        name="Failed",
        marker=dict(
            color=(100 - df["PassRate"]),
            colorscale=[
                [0.0, "#FFCCCC"],   # light red
                [0.5, "#FF6666"],   # medium red
                [1.0, "#990000"]    # dark red
            ],
            showscale=False
        ),
        text=df["Failed"],
        textposition="inside",
        textfont=dict(color="white", size=12)
    ))

    # --- Pass Rate Line ---
    fig.add_trace(go.Scatter(
        x=df["Dept"],
        y=df["PassRate"],
        name="Pass Rate (%)",
        yaxis="y2",
        mode="lines+markers+text",
        text=[f"{r:.1f}%" for r in df["PassRate"]],
        textposition="top center",
        textfont=dict(color="#1A8DAA", size=12, weight="bold", shadow=4),
        line=dict(color="#0DCDFD", width=1),
        marker=dict(size=8)
    ))

    # --- Add Total Labels above bars ---
    # This shows Passed / Total above the stacked bar
    # for i, row in df.iterrows():
    #     fig.add_annotation(
    #         x=row["Dept"],
    #         y=row["Total"] + 2,  # a bit above the top bar
    #         text=f"{int(row['Passed'])}/{int(row['Total'])}",
    #         showarrow=False,
    #         font=dict(color="black", size=13, family="Arial Black")
    #     )

    # --- Layout ---
    fig.update_layout(
        title=f"{exam_name} Exam Results by Department",
        xaxis=dict(title="Department"),
        yaxis=dict(title="Number of Students"),
        yaxis2=dict(
            title="Pass Rate (%)",
            overlaying="y",
            side="right",
            range=[0, 100]
        ),
        barmode="stack",
        legend=dict(x=0.02, y=0.98),
        plot_bgcolor="white",
        margin=dict(l=40, r=40, t=50, b=40),
        font=dict(family="Arial", size=12)
    )

    # Save chart as PNG using Kaleido
    chart_path = "apis/Telegram/Exam/assets/exam_summary_chart.png"
    fig.write_image(chart_path, width=1000, height=600)

    return chart_path
    
async def report_score(text):
    exam_rec = await sync_to_async(get_db_exam)()
    examid = exam_rec["ID"]
    processed = []
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    
    extracted_data = []
    response_arr = []
    
    for line in lines:
        # ✅ Skip department header lines
        if "---Department:" in line:
            response_arr.append(f"<b>Score Report: <{line}/b>")
        
        # Match pattern: extract student name, EVID, and Score
        # Example line: 🟢G21 Abel B500 (500) - 100
        match = re.search(r'([A-Za-z]+)\s.*\((\d+)\)\s*-\s*(\d+)', line)
        if match:
            student_name = match.group(1)
            EVID = int(match.group(2))
            Score = int(match.group(3))
            if Score > 0:
                extracted_data.append((student_name, EVID, Score))
    
    if not extracted_data:
        return
    
    # Build the reply
    response_lines = []
    processed = await sync_to_async(save_scores)(extracted_data, examid)
    
    response_lines = [f"✅ <b>{name}</b> (<b>{evid}</b>) — Score: <b>{score}</b>" for name, evid, score, _ in processed]
    
    response_arr.extend(response_lines)
    response = "\n".join(response_arr)
    return response
    
    
def save_scores(data, examid):
    reporter = "A006Z"
    local_processed = []
    for name, evid, score in data:
        with connection.cursor() as cursor:
            cursor.execute(
                "EXEC spExamReportScore @ExamID=%s, @EVID=%s, @Score=%s, @Reporter=%s",
                [examid, evid, score, reporter]
            )
            try:
                result = [dict(zip([col[0] for col in cursor.description], rec)) for rec in cursor.fetchall()]
            except Exception:
                result = []
        local_processed.append((name, evid, score, result))
    return local_processed